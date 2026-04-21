import argparse
import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REVENUE_MODELS = {
    "volume_price": ["volume", "price"],
    "users_arpu": ["users", "arpu"],
    "orders_aov": ["orders", "aov"],
    "gmv_take_rate": ["gmv", "take_rate"],
    "stores_sales_per_store": ["avg_stores", "sales_per_store"],
    "backlog_recognition": ["opening_backlog", "recognition_rate"],
    "capacity_utilization_price": ["capacity", "utilization", "price"],
}

COST_MODELS = {"gross_margin", "cogs_ratio", "unit_cost"}
COMPANY_ASSUMPTION_KEYS = [
    "sga_ratio",
    "rnd_ratio",
    "other_opex_ratio",
    "non_operating",
    "tax_rate",
    "diluted_shares",
]


def parse_args():
    parser = argparse.ArgumentParser(description="Build an investment-banking-style Excel model workbook.")
    parser.add_argument("--input", required=True, help="Path to the input JSON.")
    parser.add_argument("--output", required=True, help="Path to the output XLSX.")
    return parser.parse_args()


def load_payload(path):
    with open(path, "r", encoding="utf-8") as handle:
        payload = json.load(handle)
    validate_payload(payload)
    return payload


def validate_payload(payload):
    quarters = payload.get("quarters") or []
    if not quarters:
        raise ValueError("quarters is required")

    company = payload.get("company") or {}
    for key in ["name", "ticker", "currency", "unit"]:
        if key not in company:
            raise ValueError(f"company.{key} is required")

    forecast_start = payload.get("forecast_start")
    if forecast_start and forecast_start not in quarters:
        raise ValueError("forecast_start must be one of the quarter labels")

    segments = payload.get("segments") or []
    if not segments:
        raise ValueError("segments is required")

    seen_names = set()
    for segment in segments:
        name = segment.get("name")
        if not name:
            raise ValueError("each segment must have a name")
        if name in seen_names:
            raise ValueError(f"duplicate segment name: {name}")
        seen_names.add(name)

        revenue_model = segment.get("revenue_model")
        if revenue_model not in REVENUE_MODELS:
            raise ValueError(f"unsupported revenue_model for {name}: {revenue_model}")

        drivers = segment.get("drivers") or {}
        for driver_key in REVENUE_MODELS[revenue_model]:
            validate_series_node(drivers.get(driver_key), len(quarters), f"{name}.{driver_key}")

        cost_model = segment.get("cost_model")
        if cost_model not in COST_MODELS:
            raise ValueError(f"unsupported cost_model for {name}: {cost_model}")

        cost_assumptions = segment.get("cost_assumptions") or {}
        if cost_model == "gross_margin":
            validate_series_node(cost_assumptions.get("gross_margin"), len(quarters), f"{name}.gross_margin")
        elif cost_model == "cogs_ratio":
            validate_series_node(cost_assumptions.get("cogs_ratio"), len(quarters), f"{name}.cogs_ratio")
        else:
            validate_series_node(cost_assumptions.get("unit_cost"), len(quarters), f"{name}.unit_cost")
            volume_driver = segment.get("cost_volume_driver")
            if volume_driver not in drivers:
                raise ValueError(f"{name}.cost_volume_driver must point to a revenue driver")

    company_assumptions = payload.get("company_assumptions") or {}
    for key in COMPANY_ASSUMPTION_KEYS:
        validate_series_node(company_assumptions.get(key), len(quarters), f"company_assumptions.{key}")

    for benchmark in payload.get("benchmarks", []):
        for key in ["type", "entity", "metric", "reference_value", "source", "date", "note"]:
            if key not in benchmark:
                raise ValueError(f"benchmark rows require {key}")


def validate_series_node(node, expected_length, label):
    if not isinstance(node, dict):
        raise ValueError(f"{label} must be an object")
    values = node.get("values")
    if not isinstance(values, list) or len(values) != expected_length:
        raise ValueError(f"{label}.values must match the quarter count")


def year_label(quarter_label):
    match = re.match(r"(\d{4})", quarter_label)
    if not match:
        raise ValueError(f"quarter labels must start with the year: {quarter_label}")
    return f"{match.group(1)}{'E' if quarter_label.endswith('E') else 'A'}"


def cell_ref(sheet_name, row, col):
    return f"'{sheet_name}'!{get_column_letter(col)}{row}"


def series_number_format(unit):
    if not unit:
        return "#,##0.0"
    normalized = unit.lower()
    if "%" in unit or "ratio" in normalized or "margin" in normalized or normalized == "pct":
        return "0.0%"
    return "#,##0.0"


def apply_series_style(cell, kind, number_format):
    color_map = {
        "input": "0000FF",
        "link": "008000",
        "formula": "000000",
        "text": "000000",
    }
    cell.font = Font(color=color_map.get(kind, "000000"))
    if kind == "input":
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
    cell.number_format = number_format


def write_header(ws, title, headers, freeze_col=5):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    fill = PatternFill("solid", fgColor="D9EAF7")
    border = Border(bottom=Side(style="thin", color="5B9BD5"))
    for col_idx, value in enumerate(headers, start=1):
        cell = ws.cell(2, col_idx, value)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = f"{get_column_letter(freeze_col)}3"
    ws.sheet_view.showGridLines = False


def auto_width(ws):
    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        width = 10
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 2, 36))
        ws.column_dimensions[letter].width = width


def write_series_row(ws, row, meta_values, series_values, start_col, kind, unit):
    for idx, value in enumerate(meta_values, start=1):
        ws.cell(row, idx, value)
    for offset, value in enumerate(series_values):
        cell = ws.cell(row, start_col + offset, value)
        apply_series_style(cell, kind, series_number_format(unit))


def set_formula(ws, row, col, formula, unit, kind="formula"):
    cell = ws.cell(row, col, formula)
    apply_series_style(cell, kind, series_number_format(unit))


def set_value(ws, row, col, value, unit, kind="formula"):
    cell = ws.cell(row, col, value)
    apply_series_style(cell, kind, series_number_format(unit))


def build_assumptions_sheet(wb, payload):
    ws = wb.active
    ws.title = "Assumptions"
    quarters = payload["quarters"]
    write_header(
        ws,
        "Assumptions",
        ["Section", "Entity", "Metric", "Unit", "Benchmark", "Rationale"] + quarters,
        freeze_col=7,
    )
    quarter_start_col = 7
    row = 3
    assumption_rows = {}

    for segment in payload["segments"]:
        for key, node in segment["drivers"].items():
            write_series_row(
                ws,
                row,
                ["Revenue Driver", segment["name"], key, node.get("unit", ""), node.get("benchmark", ""), node.get("rationale", "")],
                node["values"],
                quarter_start_col,
                "input",
                node.get("unit", ""),
            )
            assumption_rows[("segment", segment["name"], key)] = row
            row += 1

        for key, node in segment["cost_assumptions"].items():
            write_series_row(
                ws,
                row,
                ["Cost Driver", segment["name"], key, node.get("unit", ""), node.get("benchmark", ""), node.get("rationale", "")],
                node["values"],
                quarter_start_col,
                "input",
                node.get("unit", ""),
            )
            assumption_rows[("segment", segment["name"], key)] = row
            row += 1

    for key, node in payload["company_assumptions"].items():
        write_series_row(
            ws,
            row,
            ["Company Assumption", "Corporate", key, node.get("unit", ""), node.get("benchmark", ""), node.get("rationale", "")],
            node["values"],
            quarter_start_col,
            "input",
            node.get("unit", ""),
        )
        assumption_rows[("company", "Corporate", key)] = row
        row += 1

    auto_width(ws)
    return assumption_rows, quarter_start_col


def build_benchmarks_sheet(wb, payload, assumption_rows, assumption_col_start):
    ws = wb.create_sheet("Benchmarks")
    write_header(
        ws,
        "Benchmarks",
        ["Type", "Entity", "Metric", "Reference Value", "Source", "Date", "Note", "Model Link"],
    )
    row = 3

    for item in payload.get("benchmarks", []):
        for col_idx, key in enumerate(["type", "entity", "metric", "reference_value", "source", "date", "note"], start=1):
            ws.cell(row, col_idx, item[key])
        row += 1

    for segment in payload["segments"]:
        for key, node in {**segment["drivers"], **segment["cost_assumptions"]}.items():
            ws.cell(row, 1, "assumption")
            ws.cell(row, 2, segment["name"])
            ws.cell(row, 3, key)
            ws.cell(row, 4, node.get("benchmark", ""))
            ws.cell(row, 5, "Seeded from model input")
            ws.cell(row, 6, "n/a")
            ws.cell(row, 7, node.get("rationale", ""))
            ws.cell(row, 8, f"Assumptions!{get_column_letter(assumption_col_start)}{assumption_rows[('segment', segment['name'], key)]}")
            row += 1

    for key, node in payload["company_assumptions"].items():
        ws.cell(row, 1, "assumption")
        ws.cell(row, 2, "Corporate")
        ws.cell(row, 3, key)
        ws.cell(row, 4, node.get("benchmark", ""))
        ws.cell(row, 5, "Seeded from model input")
        ws.cell(row, 6, "n/a")
        ws.cell(row, 7, node.get("rationale", ""))
        ws.cell(row, 8, f"Assumptions!{get_column_letter(assumption_col_start)}{assumption_rows[('company', 'Corporate', key)]}")
        row += 1

    auto_width(ws)


def build_revenue_formula(model, refs):
    if model == "volume_price":
        return f"={refs['volume']}*{refs['price']}"
    if model == "users_arpu":
        return f"={refs['users']}*{refs['arpu']}"
    if model == "orders_aov":
        return f"={refs['orders']}*{refs['aov']}"
    if model == "gmv_take_rate":
        return f"={refs['gmv']}*{refs['take_rate']}"
    if model == "stores_sales_per_store":
        return f"={refs['avg_stores']}*{refs['sales_per_store']}"
    if model == "backlog_recognition":
        return f"={refs['opening_backlog']}*{refs['recognition_rate']}"
    if model == "capacity_utilization_price":
        return f"={refs['capacity']}*{refs['utilization']}*{refs['price']}"
    raise ValueError(f"unsupported revenue model: {model}")


def parse_segment_rollups(segment_name):
    if " - " not in segment_name:
        return None, None
    left, right = segment_name.split(" - ", 1)
    return left.strip(), right.strip()


def grouped_sum_formula(data_col, rows):
    refs = [f"{get_column_letter(data_col)}{row}" for row in rows]
    return f"=SUM({','.join(refs)})"


def build_revenue_sheet(wb, payload, assumption_rows, assumption_col_start):
    ws = wb.create_sheet("Revenue Build")
    quarters = payload["quarters"]
    data_start_col = 5
    write_header(ws, "Revenue Build", ["Section", "Entity", "Metric", "Unit"] + quarters)
    row = 3
    summary_rows = defaultdict(list)
    segment_rollups = []

    for segment in payload["segments"]:
        local_rows = {}

        for driver_key in REVENUE_MODELS[segment["revenue_model"]]:
            local_rows[driver_key] = row
            unit = segment["drivers"][driver_key].get("unit", "")
            ws.cell(row, 1, "Revenue Driver Link")
            ws.cell(row, 2, segment["name"])
            ws.cell(row, 3, driver_key)
            ws.cell(row, 4, unit)
            for idx in range(len(quarters)):
                set_formula(
                    ws,
                    row,
                    data_start_col + idx,
                    f"={cell_ref('Assumptions', assumption_rows[('segment', segment['name'], driver_key)], assumption_col_start + idx)}",
                    unit,
                    "link",
                )
            row += 1

        revenue_row = row
        ws.cell(row, 1, "Revenue")
        ws.cell(row, 2, segment["name"])
        ws.cell(row, 3, "Segment Revenue")
        ws.cell(row, 4, payload["company"]["unit"])
        for idx in range(len(quarters)):
            refs = {
                key: f"{get_column_letter(data_start_col + idx)}{local_rows[key]}"
                for key in REVENUE_MODELS[segment["revenue_model"]]
            }
            set_formula(
                ws,
                row,
                data_start_col + idx,
                build_revenue_formula(segment["revenue_model"], refs),
                payload["company"]["unit"],
            )
        row += 1

        if segment["cost_model"] == "gross_margin":
            cost_key = "gross_margin"
        elif segment["cost_model"] == "cogs_ratio":
            cost_key = "cogs_ratio"
        else:
            cost_key = "unit_cost"

        local_rows[cost_key] = row
        cost_unit = segment["cost_assumptions"][cost_key].get("unit", "")
        ws.cell(row, 1, "Cost Driver Link")
        ws.cell(row, 2, segment["name"])
        ws.cell(row, 3, cost_key)
        ws.cell(row, 4, cost_unit)
        for idx in range(len(quarters)):
            set_formula(
                ws,
                row,
                data_start_col + idx,
                f"={cell_ref('Assumptions', assumption_rows[('segment', segment['name'], cost_key)], assumption_col_start + idx)}",
                cost_unit,
                "link",
            )
        row += 1

        cogs_row = row
        ws.cell(row, 1, "COGS")
        ws.cell(row, 2, segment["name"])
        ws.cell(row, 3, "Segment COGS")
        ws.cell(row, 4, payload["company"]["unit"])
        for idx in range(len(quarters)):
            revenue_ref = f"{get_column_letter(data_start_col + idx)}{revenue_row}"
            if segment["cost_model"] == "gross_margin":
                gm_ref = f"{get_column_letter(data_start_col + idx)}{local_rows['gross_margin']}"
                formula = f"={revenue_ref}*(1-{gm_ref})"
            elif segment["cost_model"] == "cogs_ratio":
                ratio_ref = f"{get_column_letter(data_start_col + idx)}{local_rows['cogs_ratio']}"
                formula = f"={revenue_ref}*{ratio_ref}"
            else:
                volume_key = segment["cost_volume_driver"]
                volume_ref = f"{get_column_letter(data_start_col + idx)}{local_rows[volume_key]}"
                unit_cost_ref = f"{get_column_letter(data_start_col + idx)}{local_rows['unit_cost']}"
                formula = f"={volume_ref}*{unit_cost_ref}"
            set_formula(ws, row, data_start_col + idx, formula, payload["company"]["unit"])
        row += 1

        gp_row = row
        ws.cell(row, 1, "Gross Profit")
        ws.cell(row, 2, segment["name"])
        ws.cell(row, 3, "Segment Gross Profit")
        ws.cell(row, 4, payload["company"]["unit"])
        for idx in range(len(quarters)):
            revenue_ref = f"{get_column_letter(data_start_col + idx)}{revenue_row}"
            cogs_ref = f"{get_column_letter(data_start_col + idx)}{cogs_row}"
            set_formula(ws, row, data_start_col + idx, f"={revenue_ref}-{cogs_ref}", payload["company"]["unit"])
        row += 2

        summary_rows["revenue"].append(revenue_row)
        summary_rows["cogs"].append(cogs_row)
        summary_rows["gross_profit"].append(gp_row)
        ip_name, type_name = parse_segment_rollups(segment["name"])
        segment_rollups.append(
            {
                "segment": segment["name"],
                "ip": ip_name,
                "type": type_name,
                "revenue_row": revenue_row,
                "cogs_row": cogs_row,
                "gross_profit_row": gp_row,
            }
        )

    ip_groups = defaultdict(lambda: defaultdict(list))
    type_groups = defaultdict(lambda: defaultdict(list))
    for item in segment_rollups:
        if item["ip"]:
            ip_groups[item["ip"]]["revenue"].append(item["revenue_row"])
            ip_groups[item["ip"]]["cogs"].append(item["cogs_row"])
            ip_groups[item["ip"]]["gross_profit"].append(item["gross_profit_row"])
        if item["type"]:
            type_groups[item["type"]]["revenue"].append(item["revenue_row"])
            type_groups[item["type"]]["cogs"].append(item["cogs_row"])
            type_groups[item["type"]]["gross_profit"].append(item["gross_profit_row"])

    for group_name, groups, section_label in [
        ("IP", ip_groups, "IP Summary"),
        ("Type", type_groups, "Type Summary"),
    ]:
        if not groups:
            continue
        row += 1
        for entity in sorted(groups):
            for metric_label, key in [
                (f"{group_name} Revenue", "revenue"),
                (f"{group_name} COGS", "cogs"),
                (f"{group_name} Gross Profit", "gross_profit"),
            ]:
                ws.cell(row, 1, section_label)
                ws.cell(row, 2, entity)
                ws.cell(row, 3, metric_label)
                ws.cell(row, 4, payload["company"]["unit"])
                for idx in range(len(quarters)):
                    data_col = data_start_col + idx
                    set_formula(
                        ws,
                        row,
                        data_col,
                        grouped_sum_formula(data_col, groups[entity][key]),
                        payload["company"]["unit"],
                    )
                row += 1

    total_rows = {}
    for label, key in [("Total Revenue", "revenue"), ("Total COGS", "cogs"), ("Total Gross Profit", "gross_profit")]:
        total_rows[key] = row
        ws.cell(row, 1, "Summary")
        ws.cell(row, 2, "Corporate")
        ws.cell(row, 3, label)
        ws.cell(row, 4, payload["company"]["unit"])
        for idx in range(len(quarters)):
            refs = [f"{get_column_letter(data_start_col + idx)}{source_row}" for source_row in summary_rows[key]]
            set_formula(ws, row, data_start_col + idx, f"=SUM({','.join(refs)})", payload["company"]["unit"])
        row += 1

    auto_width(ws)
    return total_rows, data_start_col


def build_quarterly_sheet(wb, payload, revenue_total_rows, revenue_data_start_col, assumption_rows, assumption_col_start):
    ws = wb.create_sheet("Quarterly Model")
    quarters = payload["quarters"]
    data_start_col = 4
    write_header(ws, "Quarterly Model", ["Metric", "Unit", "Note"] + quarters)

    rows = [
        ("Revenue", payload["company"]["unit"], "Link from revenue build"),
        ("YoY Growth", "%", "vs same quarter prior year"),
        ("COGS", payload["company"]["unit"], "Link from revenue build"),
        ("Gross Profit", payload["company"]["unit"], "Revenue less COGS"),
        ("Gross Margin", "%", "Gross profit divided by revenue"),
        ("SG&A Ratio", "%", "Editable assumption"),
        ("SG&A", payload["company"]["unit"], "Revenue times SG&A ratio"),
        ("R&D Ratio", "%", "Editable assumption"),
        ("R&D", payload["company"]["unit"], "Revenue times R&D ratio"),
        ("Other Opex Ratio", "%", "Editable assumption"),
        ("Other Opex", payload["company"]["unit"], "Revenue times other opex ratio"),
        ("EBIT", payload["company"]["unit"], "Gross profit less opex"),
        ("EBIT Margin", "%", "EBIT divided by revenue"),
        ("Non-operating", payload["company"]["unit"], "Editable assumption"),
        ("Pre-tax Profit", payload["company"]["unit"], "EBIT plus non-operating"),
        ("Tax Rate", "%", "Editable assumption"),
        ("Tax", payload["company"]["unit"], "Tax on positive pre-tax profit"),
        ("Net Profit", payload["company"]["unit"], "Pre-tax profit less tax"),
        ("Diluted Shares", "mn shares", "Editable assumption"),
        ("EPS", payload["company"]["unit"], "Net profit divided by diluted shares"),
    ]

    row_map = {}
    row = 3
    for metric, unit, note in rows:
        row_map[metric] = row
        ws.cell(row, 1, metric)
        ws.cell(row, 2, unit)
        ws.cell(row, 3, note)
        row += 1

    company_map = {
        "SG&A Ratio": "sga_ratio",
        "R&D Ratio": "rnd_ratio",
        "Other Opex Ratio": "other_opex_ratio",
        "Non-operating": "non_operating",
        "Tax Rate": "tax_rate",
        "Diluted Shares": "diluted_shares",
    }

    for idx in range(len(quarters)):
        q_col = data_start_col + idx
        source_col = revenue_data_start_col + idx

        set_formula(ws, row_map["Revenue"], q_col, f"={cell_ref('Revenue Build', revenue_total_rows['revenue'], source_col)}", payload["company"]["unit"], "link")
        set_formula(ws, row_map["COGS"], q_col, f"={cell_ref('Revenue Build', revenue_total_rows['cogs'], source_col)}", payload["company"]["unit"], "link")
        set_formula(ws, row_map["Gross Profit"], q_col, f"={get_column_letter(q_col)}{row_map['Revenue']}-{get_column_letter(q_col)}{row_map['COGS']}", payload["company"]["unit"])
        set_formula(ws, row_map["Gross Margin"], q_col, f"=IF({get_column_letter(q_col)}{row_map['Revenue']}=0,0,{get_column_letter(q_col)}{row_map['Gross Profit']}/{get_column_letter(q_col)}{row_map['Revenue']})", "%")

        if idx >= 4:
            current_ref = f"{get_column_letter(q_col)}{row_map['Revenue']}"
            prior_ref = f"{get_column_letter(q_col - 4)}{row_map['Revenue']}"
            set_formula(ws, row_map["YoY Growth"], q_col, f"=IF({prior_ref}=0,0,{current_ref}/{prior_ref}-1)", "%")
        else:
            set_value(ws, row_map["YoY Growth"], q_col, "", "%")

        for metric, assumption_key in company_map.items():
            unit = ws.cell(row_map[metric], 2).value
            set_formula(
                ws,
                row_map[metric],
                q_col,
                f"={cell_ref('Assumptions', assumption_rows[('company', 'Corporate', assumption_key)], assumption_col_start + idx)}",
                unit,
                "link",
            )

        set_formula(ws, row_map["SG&A"], q_col, f"={get_column_letter(q_col)}{row_map['Revenue']}*{get_column_letter(q_col)}{row_map['SG&A Ratio']}", payload["company"]["unit"])
        set_formula(ws, row_map["R&D"], q_col, f"={get_column_letter(q_col)}{row_map['Revenue']}*{get_column_letter(q_col)}{row_map['R&D Ratio']}", payload["company"]["unit"])
        set_formula(ws, row_map["Other Opex"], q_col, f"={get_column_letter(q_col)}{row_map['Revenue']}*{get_column_letter(q_col)}{row_map['Other Opex Ratio']}", payload["company"]["unit"])
        set_formula(ws, row_map["EBIT"], q_col, f"={get_column_letter(q_col)}{row_map['Gross Profit']}-{get_column_letter(q_col)}{row_map['SG&A']}-{get_column_letter(q_col)}{row_map['R&D']}-{get_column_letter(q_col)}{row_map['Other Opex']}", payload["company"]["unit"])
        set_formula(ws, row_map["EBIT Margin"], q_col, f"=IF({get_column_letter(q_col)}{row_map['Revenue']}=0,0,{get_column_letter(q_col)}{row_map['EBIT']}/{get_column_letter(q_col)}{row_map['Revenue']})", "%")
        set_formula(ws, row_map["Pre-tax Profit"], q_col, f"={get_column_letter(q_col)}{row_map['EBIT']}+{get_column_letter(q_col)}{row_map['Non-operating']}", payload["company"]["unit"])
        set_formula(ws, row_map["Tax"], q_col, f"=MAX(0,{get_column_letter(q_col)}{row_map['Pre-tax Profit']}*{get_column_letter(q_col)}{row_map['Tax Rate']})", payload["company"]["unit"])
        set_formula(ws, row_map["Net Profit"], q_col, f"={get_column_letter(q_col)}{row_map['Pre-tax Profit']}-{get_column_letter(q_col)}{row_map['Tax']}", payload["company"]["unit"])
        set_formula(ws, row_map["EPS"], q_col, f"=IF({get_column_letter(q_col)}{row_map['Diluted Shares']}=0,0,{get_column_letter(q_col)}{row_map['Net Profit']}/{get_column_letter(q_col)}{row_map['Diluted Shares']})", payload["company"]["unit"])

    auto_width(ws)
    return row_map, data_start_col


def build_annual_sheet(wb, payload, quarterly_row_map, quarterly_data_start_col):
    ws = wb.create_sheet("Annual Model")
    quarter_years = [year_label(q) for q in payload["quarters"]]
    year_indices = defaultdict(list)
    ordered_years = []
    for idx, year in enumerate(quarter_years):
        if year not in year_indices:
            ordered_years.append(year)
        year_indices[year].append(idx)

    data_start_col = 4
    write_header(ws, "Annual Model", ["Metric", "Unit", "Note"] + ordered_years)

    rows = [
        ("Revenue", payload["company"]["unit"], "Sum of quarterly revenue"),
        ("YoY Growth", "%", "vs prior year"),
        ("COGS", payload["company"]["unit"], "Sum of quarterly COGS"),
        ("Gross Profit", payload["company"]["unit"], "Revenue less COGS"),
        ("Gross Margin", "%", "Gross profit divided by revenue"),
        ("SG&A", payload["company"]["unit"], "Sum of quarterly SG&A"),
        ("R&D", payload["company"]["unit"], "Sum of quarterly R&D"),
        ("Other Opex", payload["company"]["unit"], "Sum of quarterly other opex"),
        ("EBIT", payload["company"]["unit"], "Gross profit less opex"),
        ("EBIT Margin", "%", "EBIT divided by revenue"),
        ("Non-operating", payload["company"]["unit"], "Sum of quarterly non-operating"),
        ("Pre-tax Profit", payload["company"]["unit"], "EBIT plus non-operating"),
        ("Tax", payload["company"]["unit"], "Sum of quarterly tax"),
        ("Net Profit", payload["company"]["unit"], "Pre-tax profit less tax"),
        ("Diluted Shares", "mn shares", "Average quarterly diluted shares"),
        ("EPS", payload["company"]["unit"], "Net profit divided by diluted shares"),
    ]

    row_map = {}
    row = 3
    for metric, unit, note in rows:
        row_map[metric] = row
        ws.cell(row, 1, metric)
        ws.cell(row, 2, unit)
        ws.cell(row, 3, note)
        row += 1

    sum_metrics = ["Revenue", "COGS", "Gross Profit", "SG&A", "R&D", "Other Opex", "EBIT", "Non-operating", "Pre-tax Profit", "Tax", "Net Profit"]
    for year_offset, year in enumerate(ordered_years):
        year_col = data_start_col + year_offset
        q_cols = [quarterly_data_start_col + idx for idx in year_indices[year]]

        for metric in sum_metrics:
            refs = [f"'Quarterly Model'!{get_column_letter(col)}{quarterly_row_map[metric]}" for col in q_cols]
            set_formula(ws, row_map[metric], year_col, f"=SUM({','.join(refs)})", payload["company"]["unit"])

        share_refs = [f"'Quarterly Model'!{get_column_letter(col)}{quarterly_row_map['Diluted Shares']}" for col in q_cols]
        set_formula(ws, row_map["Diluted Shares"], year_col, f"=AVERAGE({','.join(share_refs)})", "mn shares")
        set_formula(ws, row_map["Gross Margin"], year_col, f"=IF({get_column_letter(year_col)}{row_map['Revenue']}=0,0,{get_column_letter(year_col)}{row_map['Gross Profit']}/{get_column_letter(year_col)}{row_map['Revenue']})", "%")
        set_formula(ws, row_map["EBIT Margin"], year_col, f"=IF({get_column_letter(year_col)}{row_map['Revenue']}=0,0,{get_column_letter(year_col)}{row_map['EBIT']}/{get_column_letter(year_col)}{row_map['Revenue']})", "%")
        set_formula(ws, row_map["EPS"], year_col, f"=IF({get_column_letter(year_col)}{row_map['Diluted Shares']}=0,0,{get_column_letter(year_col)}{row_map['Net Profit']}/{get_column_letter(year_col)}{row_map['Diluted Shares']})", payload["company"]["unit"])

        if year_offset > 0:
            curr_ref = f"{get_column_letter(year_col)}{row_map['Revenue']}"
            prev_ref = f"{get_column_letter(year_col - 1)}{row_map['Revenue']}"
            set_formula(ws, row_map["YoY Growth"], year_col, f"=IF({prev_ref}=0,0,{curr_ref}/{prev_ref}-1)", "%")
        else:
            set_value(ws, row_map["YoY Growth"], year_col, "", "%")

    auto_width(ws)
    return row_map, data_start_col, ordered_years, year_indices


def add_check_row(ws, row, check_name, period, formula, tolerance=0.01):
    ws.cell(row, 1, check_name)
    ws.cell(row, 2, period)
    ws.cell(row, 3, formula)
    result_cell = ws.cell(row, 4, formula)
    result_cell.number_format = "#,##0.000"
    ws.cell(row, 5, f'=IF(ABS(D{row})<{tolerance},"OK","FLAG")')
    return row + 1


def build_checks_sheet(wb, payload, revenue_total_rows, revenue_data_start_col, quarterly_row_map, quarterly_data_start_col, annual_row_map, annual_data_start_col, ordered_years, year_indices):
    ws = wb.create_sheet("Checks")
    write_header(ws, "Checks", ["Check", "Period", "Formula", "Result", "Status"])
    ws["F1"] = "Flag Count"
    ws["G1"] = '=COUNTIF(E:E,"FLAG")'
    ws["F1"].font = Font(bold=True)
    ws["G1"].font = Font(bold=True)
    row = 3

    for idx, quarter in enumerate(payload["quarters"]):
        q_col = quarterly_data_start_col + idx
        revenue_col = revenue_data_start_col + idx
        row = add_check_row(
            ws,
            row,
            "Revenue build to P&L",
            quarter,
            f"='Quarterly Model'!{get_column_letter(q_col)}{quarterly_row_map['Revenue']}-'Revenue Build'!{get_column_letter(revenue_col)}{revenue_total_rows['revenue']}",
        )
        row = add_check_row(
            ws,
            row,
            "Gross profit build to P&L",
            quarter,
            f"='Quarterly Model'!{get_column_letter(q_col)}{quarterly_row_map['Gross Profit']}-'Revenue Build'!{get_column_letter(revenue_col)}{revenue_total_rows['gross_profit']}",
        )
        row = add_check_row(
            ws,
            row,
            "Pre-tax bridge",
            quarter,
            f"='Quarterly Model'!{get_column_letter(q_col)}{quarterly_row_map['Pre-tax Profit']}-('Quarterly Model'!{get_column_letter(q_col)}{quarterly_row_map['EBIT']}+'Quarterly Model'!{get_column_letter(q_col)}{quarterly_row_map['Non-operating']})",
        )
        row = add_check_row(
            ws,
            row,
            "Net profit bridge",
            quarter,
            f"='Quarterly Model'!{get_column_letter(q_col)}{quarterly_row_map['Net Profit']}-('Quarterly Model'!{get_column_letter(q_col)}{quarterly_row_map['Pre-tax Profit']}-'Quarterly Model'!{get_column_letter(q_col)}{quarterly_row_map['Tax']})",
        )

    for year_offset, year in enumerate(ordered_years):
        year_col = annual_data_start_col + year_offset
        q_cols = [quarterly_data_start_col + idx for idx in year_indices[year]]
        revenue_refs = ",".join(f"'Quarterly Model'!{get_column_letter(col)}{quarterly_row_map['Revenue']}" for col in q_cols)
        net_refs = ",".join(f"'Quarterly Model'!{get_column_letter(col)}{quarterly_row_map['Net Profit']}" for col in q_cols)
        row = add_check_row(
            ws,
            row,
            "Annual revenue roll-up",
            year,
            f"='Annual Model'!{get_column_letter(year_col)}{annual_row_map['Revenue']}-SUM({revenue_refs})",
        )
        row = add_check_row(
            ws,
            row,
            "Annual net income roll-up",
            year,
            f"='Annual Model'!{get_column_letter(year_col)}{annual_row_map['Net Profit']}-SUM({net_refs})",
        )

    auto_width(ws)


def decorate_workbook(wb):
    for ws in wb.worksheets:
        ws.auto_filter.ref = ws.dimensions


def build_workbook(payload, output_path):
    wb = Workbook()
    assumption_rows, assumption_col_start = build_assumptions_sheet(wb, payload)
    build_benchmarks_sheet(wb, payload, assumption_rows, assumption_col_start)
    revenue_total_rows, revenue_data_start_col = build_revenue_sheet(wb, payload, assumption_rows, assumption_col_start)
    quarterly_row_map, quarterly_data_start_col = build_quarterly_sheet(
        wb,
        payload,
        revenue_total_rows,
        revenue_data_start_col,
        assumption_rows,
        assumption_col_start,
    )
    annual_row_map, annual_data_start_col, ordered_years, year_indices = build_annual_sheet(
        wb,
        payload,
        quarterly_row_map,
        quarterly_data_start_col,
    )
    build_checks_sheet(
        wb,
        payload,
        revenue_total_rows,
        revenue_data_start_col,
        quarterly_row_map,
        quarterly_data_start_col,
        annual_row_map,
        annual_data_start_col,
        ordered_years,
        year_indices,
    )
    decorate_workbook(wb)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    args = parse_args()
    payload = load_payload(args.input)
    output_path = Path(args.output)
    build_workbook(payload, output_path)
    print(f"Workbook written to {output_path}")


if __name__ == "__main__":
    main()
